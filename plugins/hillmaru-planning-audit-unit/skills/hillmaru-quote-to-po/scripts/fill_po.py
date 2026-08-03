#!/usr/bin/env python3
"""
견적서에서 추출한 품목 리스트를 힐마루 발주서 엑셀 템플릿의 품목표에 채워 넣는다.

사용법:
    python fill_po.py --template <발주서_템플릿.xlsx> --items <items.json> --output <발주서_결과.xlsx>

items.json 형식 (리스트):
[
  {"품명": "스위치", "규격": "A46210030", "단위": "EA", "수량": 10, "단가": 27000},
  ...
]

동작 방식:
- 발주서 템플릿에서 "품명/규격/단위/수량/단가/공급가액" 헤더 행을 자동으로 찾는다.
- 헤더 바로 아래부터, 헤더와 동일한 열 병합 패턴을 가진 행들을 "품목 입력 가능 행"으로 판단한다.
  (합계/공급가액 요약 행처럼 병합 패턴이 달라지면 그 지점에서 품목 영역이 끝난 것으로 본다.)
- 공급가액(=단가*수량) 열은 기존 수식을 그대로 두고 절대 덮어쓰지 않는다. 그 외 헤더/합계/
  발주자·공급자 정보 등 다른 셀도 전혀 건드리지 않는다.
- 품목 수가 입력 가능 행 수보다 많으면 채우지 않고 에러를 낸다 (사용자에게 알려서 판단하게 함).
"""

import argparse
import json
import sys
import openpyxl


def norm(s):
    if s is None:
        return ""
    return str(s).replace(" ", "").replace("\u3000", "").strip()


HEADER_LABELS = {
    "품명": "품명",
    "규격": "규격",
    "단위": "단위",
    "수량": "수량",
    "단가": "단가",
    "공급가액": "공급가액",
}


def find_merge_top_left(ws, coord):
    """coord가 속한 병합 범위의 시작 셀 좌표(문자열)와 (min_col,max_col,min_row,max_row)를 반환.
    병합되어 있지 않으면 coord 자신과 1x1 범위를 반환."""
    cell = ws[coord]
    for mc in ws.merged_cells.ranges:
        if cell.coordinate in mc:
            return mc.min_col, mc.max_col, mc.min_row, mc.max_row
    return cell.column, cell.column, cell.row, cell.row


def find_header_row(ws, max_scan_row=40):
    """품명/규격/단위/수량/단가/공급가액 라벨이 모두 있는 행을 찾아
    {필드명: 시작열idx} 딕셔너리와 header_row 번호를 반환."""
    for r in range(1, max_scan_row + 1):
        found = {}
        for c in range(1, ws.max_column + 1):
            v = norm(ws.cell(row=r, column=c).value)
            for field, label in HEADER_LABELS.items():
                if v == label and field not in found:
                    found[field] = c
        if len(found) == len(HEADER_LABELS):
            return r, found
    raise ValueError(
        "발주서 템플릿에서 '품명/규격/단위/수량/단가/공급가액' 헤더 행을 찾지 못했습니다. "
        "템플릿 구조가 예상과 다를 수 있습니다."
    )


def get_field_span(ws, header_row, col_idx):
    """헤더 셀의 병합 범위(min_col,max_col)를 반환 (item row와 열 패턴 비교용)."""
    coord = ws.cell(row=header_row, column=col_idx).coordinate
    min_c, max_c, _, _ = find_merge_top_left(ws, coord)
    return min_c, max_c


def find_item_rows(ws, header_row, field_cols):
    """헤더 바로 아래부터, 품명 필드 기준으로 헤더와 동일한 열 병합 폭을 가지는
    연속된 행들을 품목 입력 가능 행으로 판단해서 리스트로 반환."""
    name_col = field_cols["품명"]
    name_min, name_max = get_field_span(ws, header_row, name_col)
    expected_width = name_max - name_min

    item_rows = []
    r = header_row + 1
    while r <= ws.max_row:
        coord = ws.cell(row=r, column=name_col).coordinate
        min_c, max_c, min_r, max_r = find_merge_top_left(ws, coord)
        # 이 행에서 시작하는 병합이 아니면 (즉 윗 행 병합의 연장이면) 종료
        if min_r != r:
            break
        width = max_c - min_c
        if width != expected_width:
            break
        item_rows.append(r)
        r += 1
    return item_rows


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--template", required=True)
    ap.add_argument("--items", required=True, help="items.json 경로")
    ap.add_argument("--output", required=True)
    args = ap.parse_args()

    with open(args.items, "r", encoding="utf-8") as f:
        items = json.load(f)

    wb = openpyxl.load_workbook(args.template, data_only=False)
    ws = wb.active

    header_row, field_cols = find_header_row(ws)
    item_rows = find_item_rows(ws, header_row, field_cols)

    if len(items) > len(item_rows):
        print(
            f"[오류] 품목 수({len(items)})가 발주서 템플릿의 입력 가능 행 수({len(item_rows)})보다 많습니다. "
            f"템플릿에 행을 추가하거나 발주서를 나눠야 합니다. 채우지 않고 종료합니다.",
            file=sys.stderr,
        )
        sys.exit(1)

    for row, item in zip(item_rows, items):
        ws.cell(row=row, column=field_cols["품명"]).value = item.get("품명", "")
        ws.cell(row=row, column=field_cols["규격"]).value = item.get("규격", "")
        ws.cell(row=row, column=field_cols["단위"]).value = item.get("단위", "EA")
        ws.cell(row=row, column=field_cols["수량"]).value = item.get("수량")
        ws.cell(row=row, column=field_cols["단가"]).value = item.get("단가")
        # 공급가액 열: 기존 수식이 있으면 그대로 두고, 비어있으면(템플릿에 수식이
        # 누락된 경우) 단가*수량 수식을 새로 채워서 합계에서 누락되지 않게 한다.
        sv_cell = ws.cell(row=row, column=field_cols["공급가액"])
        if sv_cell.value in (None, ""):
            price_col_letter = openpyxl.utils.get_column_letter(field_cols["단가"])
            qty_col_letter = openpyxl.utils.get_column_letter(field_cols["수량"])
            sv_cell.value = f"={price_col_letter}{row}*{qty_col_letter}{row}"

    wb.save(args.output)

    subtotal = sum((it.get("단가") or 0) * (it.get("수량") or 0) for it in items)
    vat = round(subtotal * 0.1)
    total = subtotal + vat
    print(f"[완료] {len(items)}개 품목을 {args.output} 에 채워 넣었습니다. "
          f"(사용된 행: {item_rows[0]}~{item_rows[len(items)-1]})")
    print(f"[검증용] 소계: {subtotal:,} / 부가세: {vat:,} / 합계: {total:,} "
          f"— 이 값을 견적서의 합계 금액과 비교해서 확인하세요.")


if __name__ == "__main__":
    main()
